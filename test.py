from pathlib import Path

import openvino.runtime as ov
from openvino.preprocess import PrePostProcessor
from openvino.preprocess import ColorFormat
from openvino.runtime import Layout, Type
import openpyxl
import numpy as np
import cv2 as cv 
import argparse
import os

coconame = [
    "person",         "bicycle",    "car",           "motorcycle",    "airplane",     "bus",           "train",
    "truck",          "boat",       "traffic light", "fire hydrant",  "stop sign",    "parking meter", "bench",
    "bird",           "cat",        "dog",           "horse",         "sheep",        "cow",           "elephant",
    "bear",           "zebra",      "giraffe",       "backpack",      "umbrella",     "handbag",       "tie",
    "suitcase",       "frisbee",    "skis",          "snowboard",     "sports ball",  "kite",          "baseball bat",
    "baseball glove", "skateboard", "surfboard",     "tennis racket", "bottle",       "wine glass",    "cup",
    "fork",           "knife",      "spoon",         "bowl",          "banana",       "apple",         "sandwich",
    "orange",         "broccoli",   "carrot",        "hot dog",       "pizza",        "donut",         "cake",
    "chair",          "couch",      "potted plant",  "bed",           "dining table", "toilet",        "tv",
    "laptop",         "mouse",      "remote",        "keyboard",      "cell phone",   "microwave",     "oven",
    "toaster",        "sink",       "refrigerator",  "book",          "clock",        "vase",          "scissors",
    "teddy bear",     "hair drier", "toothbrush" ]

names=["nodule", "others"]

class Yolov9:
    def __init__(self, xml_model_path="./model/yolov9-c-converted.xml", conf=0.2, nms=0.4):
        # Step 1. Initialize OpenVINO Runtime core
        core = ov.Core()
        # Step 2. Read a model
        model = core.read_model(str(Path(xml_model_path)))

        # Step 3. Inizialize Preprocessing for the model
        ppp = PrePostProcessor(model)
        # Specify input image format
        ppp.input().tensor().set_element_type(Type.u8).set_layout(Layout("NHWC")).set_color_format(ColorFormat.BGR)
        #  Specify preprocess pipeline to input image without resizing
        ppp.input().preprocess().convert_element_type(Type.f32).convert_color(ColorFormat.RGB).scale([255., 255., 255.])
        # Specify model's input layout
        ppp.input().model().set_layout(Layout("NCHW"))
        #  Specify output results format

        ppp.output().tensor().set_element_type(Type.f32)
        # Embed above steps in the graph
        model = ppp.build()

        self.compiled_model = core.compile_model(model, "CPU")

        self.input_width = 640
        self.input_height = 640
        self.conf_thresh = conf
        self.nms_thresh = nms
        self.colors = []

        # Create random colors
        np.random.seed(42)  # Setting seed for reproducibility

        for i in range(len(coconame)):
            color = tuple(np.random.randint(100, 256, size=3))
            self.colors.append(color)

    def resize_and_pad(self, image):

        old_size = image.shape[:2] 
        ratio = float(self.input_width/max(old_size))#fix to accept also rectangular images
        new_size = tuple([int(x*ratio) for x in old_size])

        image = cv.resize(image, (new_size[1], new_size[0]))
        
        delta_w = self.input_width - new_size[1]
        delta_h = self.input_height - new_size[0]
        
        color = [100, 100, 100]
        new_im = cv.copyMakeBorder(image, 0, delta_h, 0, delta_w, cv.BORDER_CONSTANT, value=color)
        
        return new_im, delta_w, delta_h

    def predict(self, img):

        # Step 4. Create tensor from image
        input_tensor = np.expand_dims(img, 0)

        # Step 5. Create an infer request for model inference 
        infer_request = self.compiled_model.create_infer_request()
        infer_request.infer({0: input_tensor})

        # Step 6. Retrieve inference results 
        output = infer_request.get_output_tensor()
        detections = output.data[0].T

        # Step 7. Postprocessing including NMS  
        boxes = []
        class_ids = []
        confidences = []
        for prediction in detections:
            classes_scores = prediction[4:]
            _, _, _, max_indx = cv.minMaxLoc(classes_scores)
            class_id = max_indx[1]
            if (classes_scores[class_id] > self.conf_thresh):
                confidences.append(classes_scores[class_id])
                class_ids.append(class_id)
                x, y, w, h = prediction[0].item(), prediction[1].item(), prediction[2].item(), prediction[3].item()
                xmin = x - (w / 2)
                ymin = y - (h / 2)
                box = np.array([xmin, ymin, w, h])
                boxes.append(box)

        indexes = cv.dnn.NMSBoxes(boxes, confidences, self.conf_thresh, self.nms_thresh)

        detections = []
        for i in indexes:
            j = i.item()
            detections.append({"class_index": class_ids[j], "confidence": confidences[j], "box": boxes[j]})

        return detections

    def draw(self, img, detections, dw, dh):
        # Step 8. Print results and save Figure with detections
        for detection in detections:
        
            box = detection["box"]
            classId = detection["class_index"]
            confidence = detection["confidence"]

            rx = img.shape[1] / (self.input_width - dw)
            ry = img.shape[0] / (self.input_height - dh)
            box[0] = rx * box[0]
            box[1] = ry * box[1]
            box[2] = rx * box[2]
            box[3] = ry * box[3]

            xmax = box[0] + box[2]
            ymax = box[1] + box[3]

            # Drawing detection box
            cv.rectangle(img, (int(box[0]), int(box[1])), (int(xmax), int(ymax)), tuple(map(int, self.colors[classId])), 3)

            # Detection box text
            class_string = coconame[classId] + ' ' + str(confidence)[:4]
            text_size, _ = cv.getTextSize(class_string, cv.FONT_HERSHEY_DUPLEX, 1, 2)
            text_rect = (box[0], box[1] - 40, text_size[0] + 10, text_size[1] + 20)
            cv.rectangle(img, 
                (int(text_rect[0]), int(text_rect[1])), 
                (int(text_rect[0] + text_rect[2]), int(text_rect[1] + text_rect[3])), 
                tuple(map(int, self.colors[classId])), cv.FILLED)
            cv.putText(img, class_string, (int(box[0] + 5), int(box[1] - 10)), cv.FONT_HERSHEY_DUPLEX, 1, (0, 0, 0), 2, cv.LINE_AA)

def make_parser():
    parser = argparse.ArgumentParser("onnxruntime inference")
    parser.add_argument(
        "-m",
        "--model",
        type=str,
        default="yolov9-c-converted.onnx",
        help="Input your onnx model.",
    )
    parser.add_argument(
        "-i",
        "--data_path",
        type=str,
        default='videos/palace.mp4',
        help="Path to your input image.",
    )
    parser.add_argument(
        "-s",
        "--score_thr",
        type=float,
        default=0.1,
        help="Score threshould to filter the result.",
    )
    parser.add_argument(
        "-n",
        "--nms_thr",
        type=float,
        default=0.3,
        help="NMS threshould.",
    )
    
    return parser

# Process a single image
def process_image(model, image_path):
    img = cv.imread(image_path)
    img_resized, dw, dh = model.resize_and_pad(img)
    results = model.predict(img_resized)
    model.draw(img, results, dw, dh)
    cv.imshow("result", img)
    cv.waitKey(0)
    cv.destroyAllWindows()

# Process a folder of images
def process_folder(model, folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".jpg") or filename.endswith(".png"):
            image_path = os.path.join(folder_path, filename)
            process_image(model, image_path)

# Process a video
def process_video(model, video_path):
    cap = cv.VideoCapture(video_path)
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
        img_resized, dw, dh = model.resize_and_pad(frame)
        results = model.predict(img_resized)
        model.draw(frame, results, dw, dh)
        cv.imshow("result", frame)
        if cv.waitKey(1) & 0xFF == ord('q'):
            break
    cap.release()
    cv.destroyAllWindows()


def main():
    args = make_parser().parse_args()

    # Initialize YOLOv9 model (assuming xml openvino model)
    model = Yolov9(args.model)

    if args.data_path.endswith('.jpg') or args.data_path.endswith('.png'):
        process_image(model, args.data_path)
    elif os.path.isdir(args.data_path):
        process_folder(model, args.data_path)
    elif args.data_path.endswith('.mp4'):  # Add support for other video formats
        process_video(model, args.data_path)
    else:
        print("Error: Unsupported file format")    


# python test.py --model=C:/Users/USER/Desktop/yolov9/runs/train/lung10/weights/best.xml --data_path=combine.png
def longEdgeImageProcess(image, long_edge_size = 1024, short_edge_size = 1024):

    # Calculate aspect ratio
    aspect_ratio = image.shape[1] / image.shape[0]  # width / height
    direction = ""
    print(image.shape[1], image.shape[0])
    # Determine the dimensions for resizing while maintaining aspect ratio
    if aspect_ratio >= 1: # width > height
        new_height = int(long_edge_size / aspect_ratio)
        new_width = long_edge_size 

        pad = short_edge_size - new_height
        direction = "horizontal"
        ratio = image.shape[1] / long_edge_size
    else:
        new_height = long_edge_size
        new_width = int(long_edge_size * aspect_ratio)
        print(f"new width: {new_width}")
        
        ratio = image.shape[0] / long_edge_size
        print(new_width)
        if(new_width%2 == 1):
            new_width += 1
        pad = short_edge_size - new_width

    # Resize the image
    resized_image = cv.resize(image, (new_width, new_height))
    print(resized_image.shape)
    # # Perform zero-padding on the long side
    # top_pad = 0
    inverse_flag = False
    top_pad = 0
    # print(f"pad: {pad}")
    # if aspect_ratio < 1:
    #     resized_image = np.transpose(resized_image, (1,0,2))
    #     inverse_flag = True

    print("shape:", resized_image.shape)
    if pad > 0:
        top_pad = pad // 2
        print(top_pad)
        bottom_pad = pad - top_pad
        if aspect_ratio >= 1:  
            resized_image = cv.copyMakeBorder(resized_image, top_pad, bottom_pad, 0, 0, cv.BORDER_CONSTANT, value=[0, 0, 0])     
        else:

            resized_image = cv.copyMakeBorder(resized_image, 0, 0, top_pad, bottom_pad, cv.BORDER_CONSTANT, value=[0, 0, 0])
    print(resized_image.shape)
    return resized_image, inverse_flag, direction, top_pad, ratio



def shortEdgeImageProcess(image, short_edge_size=1024, long_edge_size=640):

    # Calculate aspect ratio
    aspect_ratio = image.shape[1] / image.shape[0]  # width / height

    # Determine the dimensions for resizing while maintaining aspect ratio
    direction = ""
    if aspect_ratio >= 1:
        new_height = short_edge_size
        new_width = int(short_edge_size * aspect_ratio)
        pad = long_edge_size - new_width
        direction = "horizontal"
        ratio = image.shape[0] / short_edge_size
    else:
        new_width = short_edge_size
        new_height = int(short_edge_size / aspect_ratio)
        pad = long_edge_size - new_height
        direction = "veritical"
        ratio = image.shape[1] / short_edge_size

    # Resize the image
    resized_image = cv.resize(image, (new_width, new_height))



    # Perform zero-padding on the long side
    top_pad = 0
    inverse_flag = False
    if pad > 0:
        top_pad = pad // 2
        print(top_pad)

        bottom_pad = pad - top_pad
        if aspect_ratio >= 1:       
            resized_image = cv.copyMakeBorder(resized_image, 0, 0, top_pad, bottom_pad, cv.BORDER_CONSTANT, value=[0, 0, 0])
            resized_image = np.transpose(resized_image, (1,0,2))
            inverse_flag = True
        else:
            resized_image = cv.copyMakeBorder(resized_image, top_pad, bottom_pad, 0, 0, cv.BORDER_CONSTANT, value=[0, 0, 0])

    return resized_image, inverse_flag, direction, top_pad, ratio

#  ==========================================================================
#  ==========================================================================
#  ==========================================================================
def preprocess_warpAffine(image, dst_width = 640, dst_height = 640):
    scale = min((dst_width / image.shape[1], dst_height / image.shape[0]))
    ox = (dst_width - scale * image.shape[1]) / 2
    oy = (dst_height - scale * image.shape[0]) / 2
    M = np.array([
        [scale, 0, ox],
        [0, scale, oy]
        ], dtype=np.float32)

    img_pre = cv.warpAffine(image, M, (dst_width, dst_height), flags = cv.INTER_LINEAR,
                            borderMode=cv.BORDER_CONSTANT, borderValue=(114,114,114))

    IM = cv.invertAffineTransform(M)

    img_pre = (img_pre[...,::-1] / 255.0).astype(np.float32)
    img_pre = img_pre.transpose(2, 0, 1)[None]

    return img_pre, IM

def postprocess(pred, IM=[], conf_thres=0.4, iou_thres=0.45): #0.7

    boxes = []
    for item in pred:
        cx, cy, w, h = item[:4]
        label = item[4:].argmax()
        confidence = item[4 + label]
        if confidence < conf_thres:
            continue

        left   = cx - w * 0.5
        top    = cy - h * 0.5
        right  = cx + w * 0.5
        bottom = cy + h * 0.5
        boxes.append([left, top, right, bottom, confidence, label])

    boxes = np.array(boxes)
    if(boxes.shape[0] == 0):
        return NMS(boxes, iou_thres)
    lr = boxes[:, [0, 2]]
    tb = boxes[:, [1, 3]]
    boxes[:, [0, 2]] = IM[0][0] * lr + IM[0][2]
    boxes[:, [1, 3]] = IM[1][1] * tb + IM[1][2]
    boxes = sorted(boxes.tolist(), key=lambda x:x[4], reverse=True)

    return NMS(boxes, iou_thres)

def iou(box1, box2):

    def area_box(box):
        return (box[2] - box[0]) * (box[3] - box[1])

    left   = max(box1[0], box2[0])
    top    = max(box1[1], box2[1])
    right  = min(box1[2], box2[2])
    bottom = min(box1[3], box2[3])

    union = max((right - left), 0) * max((bottom - top), 0)
    cross = area_box(box1) + area_box(box2) - union

    if( cross == 0 or union == 0):
        return 0 
    return union / cross  

def NMS(boxes, iou_thres):

    remove_flags = [False] * len(boxes)


    keep_boxes = []
    for i, ibox in enumerate(boxes):
        if remove_flags[i]:
            continue

        keep_boxes.append(ibox)
        for j in range(i + 1, len(boxes)):
            if(remove_flags):
                continue

            jbox = boxes[j]
            if(ibox[5] != jbox[5]):
                continue

            if(iou(ibox, jbox) > iou_thres):
                remove_flags[j] = True

    return keep_boxes

def hsv2bgr(h, s, v):
    h_i = int(h * 6)
    f = h * 6 - h_i
    p = v * (1 - s)
    q = v * (1 - f * s)
    t = v * (1 - (1 - f) * s)

    r, g, b = 0, 0, 0

    if h_i == 0:
        r, g, b = v, t, p
    elif h_i == 1:
        r, g, b = q, v, p
    elif h_i == 2:
        r, g, b = p, v, t
    elif h_i == 3:
        r, g, b = p, q, v
    elif h_i == 4:
        r, g, b = t, p, v
    elif h_i == 5:
        r, g, b = v, p, q

    return int(b * 255), int(g * 255), int(r * 255)                




class result_record():

    def __init__(self):
        self.normal_positive = 0
        self.normal_negative = 0
        self.nodule_positive = 0
        self.nodule_negative = 0
        self.other_positive = 0 
        self.other_negative = 0
        self.mix_positive = 0
        self.mix_negative = 0

    def normal_predict_box(self, boxes):
        if(len(boxes) > 0):
            self.normal_negative += 1
        else:
            self.normal_positive += 1

    def abnormal_predict_box(self, boxes, target_class, target_type):
        temp = {}
        for classname in target_class:
            temp[classname] = 0


        if(len(boxes) > 0):
            for obj in boxes:
                label = int(obj[5])
                if(names[label] in target_class):
                    temp[names[label]] += 1

            if(target_type == "nodule"):
                if(temp[target_class[0]] > 0):
                    self.nodule_positive += 1
                else:
                    self.nodule_negative += 1

            elif(target_type == "others"):
                if(temp[target_class[0]] > 0):
                    self.other_positive += 1
                else:
                    self.other_negative += 1

            elif(target_type == "mix"):
                if(temp[target_class[0]] > 0 and temp[target_class[1]] > 0):
                    self.mix_positive += 1
                else:
                    self.mix_negative += 1
            else:
                print("Cannot find target_type!")


        else:
            if(target_type == "nodule"):
                self.nodule_negative += 1

            elif(target_type == "others"):
                self.other_negative += 1

            elif(target_type == "mix"):
                self.mix_negative += 1
            
            else:
                print("Cannot find target_type!")
    
    def print_results(self):
        print(f"Normal: posivitive --> {self.normal_positive}, negative --> {self.normal_negative}")   
        print(f"Nodule: posivitive --> {self.nodule_positive}, negative --> {self.nodule_negative}")
        print(f"Others: posivitive --> {self.other_positive}, negative --> {self.other_negative}")
        print(f"Mix: posivitive --> {self.mix_positive}, negative --> {self.mix_negative}")   

    def write_to_excel(self, model_name, filename = "result.xlsx"):
        # 打開 Excel 檔案
        wb = openpyxl.load_workbook(filename)
        # 選擇第一個工作表
        sheet = wb.active
        # 計算下一個空行的索引
        next_row = sheet.max_row + 1
        # 將數據寫入下一個空行
        sheet.cell(row=next_row, column=1, value=model_name)
        sheet.cell(row=next_row, column=2, value=f"{self.normal_positive} / {self.normal_negative}")
        n = self.normal_positive + self.normal_negative
        if( n > 0):
            sheet.cell(row=next_row, column=3, value = f"{self.normal_positive / n:.2f}")
        else:
            sheet.cell(row=next_row, column=3, value = 0)
        
        sheet.cell(row=next_row, column=4, value=f"{self.nodule_positive} / {self.nodule_negative}")
        n = self.nodule_positive + self.nodule_negative
        if( n > 0):
            sheet.cell(row=next_row, column=5, value = f"{self.nodule_positive / n:.2f}")
        else:
            sheet.cell(row=next_row, column=5, value = 0)

        sheet.cell(row=next_row, column=6, value=f"{self.other_positive} / {self.other_negative}")
        n = self.other_positive + self.other_negative
        if( n > 0):
            sheet.cell(row=next_row, column=7, value = f"{self.other_positive / n:.2f}")
        else:
            sheet.cell(row=next_row, column=7, value = 0)

        sheet.cell(row=next_row, column=8, value=f"{self.mix_positive} / {self.mix_negative}")
        n = self.mix_positive + self.mix_negative
        if( n > 0):
            sheet.cell(row=next_row, column=9, value = f"{self.mix_positive / n:.2f}")
        else:
            sheet.cell(row=next_row, column=9, value = 0)
        # 保存文件
        wb.save(filename)


def random_color(id):
    h_plane = (((id << 2) ^ 0x937151) % 100) / 100.0
    s_plane = (((id << 3) ^ 0x315793) % 100) / 100.0
    return hsv2bgr(h_plane, s_plane, 1)

if __name__ == "__main__":
    #main()
    modelpath = "C:/Users/USER/Desktop/yolov9/runs/train/lung2/weights/best.xml"
    modelname = "Lung2"
    # present = datetime.datetime.now()
    core = ov.Core()
    ov_model = core.read_model(modelpath)
    compiled_model = core.compile_model(ov_model, "AUTO")
    infer_request = compiled_model.create_infer_request()

    test_type_list = [ "nodule", "infiltration", "fibrosis", "mix"]
    r = result_record()


    test_root = "D:/NCKU/dataset/test"
    positive_count = 0 
    negative_count = 0
    positive_samples = []
    for test_type in test_type_list:
        test_root_dir =  os.path.join(test_root, test_type)
        for dirname in os.listdir(test_root_dir):
            folder_path = os.path.join(test_root_dir, dirname)
            test_image_path = os.path.join(folder_path, "combine.png")
            if(not os.path.exists(test_image_path)):
                stand = cv.imread(os.path.join(folder_path, "stand.png"), -1)
                soft = cv.imread(os.path.join(folder_path, "soft.png"), -1)
                bone = cv.imread(os.path.join(folder_path, "bone.png"), -1)

                exportImage = np.zeros((*stand.shape, 3))

                exportImage[:,:,0] = bone
                exportImage[:,:,1] = soft
                exportImage[:,:,2] = stand
                
                cv.imwrite(test_image_path, exportImage)
            
            #image_path = "C:/Users/USER/Desktop/NCKU/dataset/samples/00002_A056/combine.png"
            print("="*64)
            print(f"Image Path:{test_image_path}")
            image = cv.imread(test_image_path)
            img_pre, IM = preprocess_warpAffine(image)
            # image, inverse_flag, direction, top_pad, ratio = longEdgeImageProcess(image, 640, 640)
            
            # print(f"image shape: {img_pre.shape}")
            # image = np.transpose(image, (2,0,1))
            # image = image[np.newaxis,:,:,:]

            input_tensor = ov.Tensor(array=img_pre.astype(np.float32), shared_memory=False)
            infer_request.set_input_tensor(input_tensor)
            infer_request.start_async()
            infer_request.wait()

            #print( np.squeeze(infer_request.get_output_tensor(0).data)[:,0])

            detections = np.squeeze(infer_request.get_output_tensor(0).data).T


            boxes = postprocess(detections, IM)
            if(test_type == "normal"):
                r.normal_predict_box(boxes)
            else:
                if(test_type == "nodule"):
                    r.abnormal_predict_box(boxes, ["nodule"], test_type)

                elif(test_type =="infiltration" or test_type == "fibrosis"):
                    r.abnormal_predict_box(boxes, ["others"], "others")
                
                else:
                    r.abnormal_predict_box(boxes, ["nodule", "others"], "mix")

            if(len(boxes) > 0):
                positive_count += 1
                positive_samples.append(test_image_path)
            else:
                negative_count += 1

            for obj in boxes:
                left, top, right, bottom =  int(obj[0]), int(obj[1]), int(obj[2]), int(obj[3])
                confidence = obj[4]
                label = int(obj[5])
                color = random_color(label)
                cv.rectangle(image, (left, top), (right, bottom), color=color, thickness=2, lineType=cv.LINE_AA)
                caption = f"{names[label]} {confidence:.2f}"
                print(caption)
                w, h = cv.getTextSize(caption, 0, 1, 2)[0]
                cv.rectangle(image, (left - 3, top - 33), (left + w + 10, top), color, -1)
                cv.putText(image, caption, (left, top - 5), 0, 1, (0, 0, 0), 2, 16)

        #cv.imwrite("infer.png", image)
        print("="*64)
        print(f"Positive Nodule Samples: {positive_count}")
        print(f"negative Nodule Samples: {negative_count}")
    #print(positive_samples)
    r.print_results()
    r.write_to_excel(modelname)
        #np.save('positive_samples.npy', positive_samples)

    # # 示例用法
    # filename = "你的 Excel 檔案.xlsx"
    # model_name = "你的模型"
    # sensitivity = 0.85
    # specificity = 0.92

    # write_to_excel(filename, model_name, sensitivity, specificity)


    # # Step 7. Postprocessing including NMS  
    # boxes = []
    # class_ids = []
    # confidences = []
    # for prediction in detections:
    #     classes_scores = prediction[4:]
    #     _, _, _, max_indx = cv.minMaxLoc(classes_scores)
    #     class_id = max_indx[1]
    #     if (classes_scores[class_id] > 0.45):
    #         confidences.append(classes_scores[class_id])
    #         class_ids.append(class_id)
    #         x, y, w, h = prediction[0].item(), prediction[1].item(), prediction[2].item(), prediction[3].item()
    #         xmin = x - (w / 2)
    #         ymin = y - (h / 2)
    #         box = np.array([xmin, ymin, w, h])
    #         boxes.append(box)

    # indexes = cv.dnn.NMSBoxes(boxes, confidences, 0.45, 0.4)

    # detections = []
    # for i in indexes:
    #     j = i.item()
    #     detections.append({"class_index": class_ids[j], "confidence": confidences[j], "box": boxes[j]})

    # print(detections)